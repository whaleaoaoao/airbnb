import sys
import requests
import json
import openpyxl
import time
import os


def get_json(url):
    r = requests.get(url)
    r.raise_for_status()
    data = json.loads(r.text)
    return data


# p代表获取第一部分或第二部分或2+页的数据
def ratings(id):
    room_url = 'https://zh.airbnb.com/api/v2/pdp_listing_details/%d?_format=for_rooms_show' \
               '&key=d306zoyjsyarp7ifhu67rjxn52tv0t20&' % id
    room_data = get_json(room_url)
    common = room_data['pdp_listing_detail']['review_details_interface']['review_summary']
    accuracy = common[0]['value']
    communication = common[1]['value']
    cleanliness = common[2]['value']
    location = common[3]['value']
    checkin = common[4]['value']
    value = common[5]['value']
    response_time = room_data['pdp_listing_detail']['primary_host']['response_time_without_na']
    ratings_row = [accuracy, communication, cleanliness, location, checkin, value, response_time]
    return ratings_row


location = sys.argv[1]
row1 = int(sys.argv[2])
row2 = int(sys.argv[3])
os.chdir('C:\\Users\\Administrator')
wb = openpyxl.load_workbook('%s.xlsx' % location)
sheet = wb.get_active_sheet()
# 填写第一行题目
if row1 == 2:
    titles = ['accuracy', 'communication', 'cleanliness', 'location', 'checkin', 'value']
    for i in range(21, 27):
        sheet.cell(row=1, column=i).value = titles[i-21]
for i in range(row1, row2+1):
    room_id = sheet.cell(row=i, column=2).value
    ratings_data = ratings(room_id)
    for j in range(21, 27):
        sheet.cell(row=i, column=j).value = ratings_data[j - 21]
    if sheet.cell(row=i, column=9).value == True :
        continue
    else:
        sheet.cell(row=i, column=9).value = ratings_data[6]
        wb.save('%s.xlsx' % location)
print('Row %d has completed' % row2)
