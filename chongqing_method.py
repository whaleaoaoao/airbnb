import sys
import requests
import json
import openpyxl
import time


def get_json(url):
    r = requests.get(url)
    r.raise_for_status()
    data = json.loads(r.text)
    return data


# p代表获取第一部分或第二部分或2+页的数据
def row_data(url, p, i):
    if p == 1:
        a = 1
    elif p == 2:
        a = 3
    elif p == 3:
        a = 0
    page_data = get_json(url)
    common1 = page_data['explore_tabs'][0]['sections'][a]['listings'][i]['listing']
    room_id = common1['id']
    host_profile_pic = common1['user']['has_profile_pic']
    verified = page_data['explore_tabs'][0]['sections'][a]['listings'][i]['verified']['badge_secondary_text']
    instant_book = page_data['explore_tabs'][0]['sections'][a]['listings'][i]['pricing_quote']['can_instant_book']
    business_travel = common1['is_business_travel_ready']
    fully_refundable = common1['is_fully_refundable']
    pic_count = common1['picture_count']
    price = page_data['explore_tabs'][0]['sections'][a]['listings'][i]['pricing_quote']['rate_with_service_fee']['amount']

    room_url = 'https://zh.airbnb.com/api/v2/pdp_listing_details/%d?_format=for_rooms_show' \
               '&key=d306zoyjsyarp7ifhu67rjxn52tv0t20&' % room_id
    room_data = get_json(room_url)
    common = room_data['pdp_listing_detail']
    room_name = common['name']
    rating = common['star_rating']
    reviews_count = common['visible_review_count']
    superhost = common['primary_host']['is_superhost']
    host_id = common['primary_host']['id']
    response_time = common['primary_host']['response_time_without_na']
    localized_city = common['localized_city']
    neighborhood = common['location_title']
    guest_num = common['p3_event_data_logging']['person_capacity']
    room_type = common['p3_event_data_logging']['room_type']
    member_since = common['primary_host']['member_since']
    try:
        accuracy = common['review_details_interface']['review_summary'][0]['value']
        communication = common['review_details_interface']['review_summary'][1]['value']
        cleanliness = common['review_details_interface']['review_summary'][2]['value']
        location = common['review_details_interface']['review_summary'][3]['value']
        checkin = common['review_details_interface']['review_summary'][4]['value']
        value = common['review_details_interface']['review_summary'][5]['value']
    except:
        accuracy = 0
        communication = 0
        cleanliness = 0
        location = 0
        checkin = 0
        value = 0

    cancellation_url = 'https://zh.airbnb.com/api/v2/pdp_listing_booking_details?' \
                       'listing_id=%d&_format=for_web_dateless' \
                       '&key=d306zoyjsyarp7ifhu67rjxn52tv0t20&locale=zh' % room_id
    cancellation_data = get_json(cancellation_url)
    cancellation_policy = cancellation_data['pdp_listing_booking_details'][0]['cancellation_policies'][0]['title'][:4]

    row_data = [room_name, room_id, host_id, rating, reviews_count,
                superhost, host_profile_pic, verified, response_time,
                neighborhood, instant_book, business_travel, fully_refundable, cancellation_policy,
                guest_num, room_type, pic_count,
                price, localized_city, member_since,
                accuracy, communication, cleanliness, location, checkin, value]
    return row_data


location = ' '.join(sys.argv[1:])
wb = openpyxl.Workbook()
sheet = wb.create_sheet(index=0, title='Sheet1')
titles = ['room_name', 'room_id', 'host_id',
          'rating', 'reviews_count',
          'superhost', 'host_profile_pic', 'verified', 'response_time',
          'neighborhood', 'instant_book', 'business_travel', 'fully_refundable', 'cancellation_policy',
          'guest_num', 'room_type', 'pic_count',
          'price', 'localized_city', 'member_since',
          'accuracy', 'communication', 'cleanliness', 'location', 'checkin', 'value']
for i in range(1, len(titles)+1):
    sheet.cell(row=1, column=i).value = titles[i-1]

page_url = 'https://zh.airbnb.com/api/v2/explore_tabs?version=1.3.9&satori_version=1.0.12&_format=for_exp' \
           'lore_search_web&experiences_per_grid=20&items_per_grid=18&guidebooks_per_grid=20&auto_ib=true' \
           '&fetch_filters=true&has_zero_guest_treatment=true&is_guided_search=true&is_new_cards_experime' \
           'nt=true&luxury_pre_launch=false&query_understanding_enabled=false&show_groupings=true&support' \
           's_for_you_v3=true&timezone_offset=480&client_session_id=6328b000-5e27-4733-93ad-ae74cb8b15ae&' \
           'map_toggle=true&metadata_only=false&is_standard_search=true&refinement_paths%5B%5D=%2Fhomes&s' \
           'elected_tab_id=home_tab&allow_override%5B%5D=&s_tag=NrlDFIkd&screen_size=large&query=nanjing&' \
           '_intents=p1&key=d306zoyjsyarp7ifhu67rjxn52tv0t20&currency=CNY&locale=zh'
for i in range(12):
    row = row_data(page_url, 1, i)
    for c in range(1, len(row)+1):
        sheet.cell(row=i+2, column=c).value = row[c-1]
for j in range(6):
    row = row_data(page_url, 2, j)
    for c in range(1, len(row)+1):
        sheet.cell(row=j+14, column=c).value = row[c-1]
print('Page 1 has completed.')
for i in range(300):
    time.sleep(1)

for page in range(2, 18):
    print('Started searching page %d.' % page)
    page_url = 'https://zh.airbnb.com/api/v2/explore_tabs?version=1.3.9&satori_version=1.0.12&_format=for_ex' \
               'plore_search_web&experiences_per_grid=20&items_per_grid=18&guidebooks_per_grid=20&auto_ib=tru' \
               'e&fetch_filters=true&has_zero_guest_treatment=true&is_guided_search=true&is_new_cards_experim' \
               'ent=true&luxury_pre_launch=false&query_understanding_enabled=false&show_groupings=true&suppor' \
               'ts_for_you_v3=true&timezone_offset=480&client_session_id=6328b000-5e27-4733-93ad-ae74cb8b15ae' \
               '&map_toggle=true&metadata_only=false&is_standard_search=true&refinement_paths%5B%5D=%2Fhomes&' \
               'selected_tab_id=home_tab&allow_override%5B%5D=&s_tag=NrlDFIkd&section_offset=6' \
               '&items_offset=' + str(18*(page-1)) + '&screen_size=large' \
               '&query=nanjing&_intents=p1&key=d306zoyjsyarp7ifhu67rjxn52tv0t20&currency=CNY&locale=zh'
    for i in range(18):
            row = row_data(page_url, 3, i)
            for c in range(1, len(row)+1):
                sheet.cell(row=18*(page-1)+2+i, column=c).value = row[c-1]
            for t in range(30):
                time.sleep(1)
    print('Page %d has completed.' % page)
    for i in range(300):
        time.sleep(1)

wb.save('%s.xlsx' % location)
print('All completed')
