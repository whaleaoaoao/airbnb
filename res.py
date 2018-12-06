import sys
import requests
import json
import openpyxl
import time
import os


def get_json(url):
    r = requests.get(url)
    r.raise_for_status()
    data = json.loads(r.text)
    return data


def response_time(id):
    room_url = 'https://zh.airbnb.com/api/v2/pdp_listing_details/%d?_format=for_rooms_show' \
               '&key=d306zoyjsyarp7ifhu67rjxn52tv0t20&' % id
    room_data = get_json(room_url)
    response_time = room_data['pdp_listing_detail']['primary_host']['response_time_without_na']
    return response_time


location = sys.argv[1]
row1 = int(sys.argv[2])
row2 = int(sys.argv[3])
os.chdir('C:\\Users\\Administrator')
wb = openpyxl.load_workbook('%s.xlsx' % location)
sheet = wb.get_active_sheet()
for i in range(row1, row2+1):
    if sheet.cell(row=i, column=9).value == True :
        continue
    else:
        room_id = sheet.cell(row=i, column=2).value
        sheet.cell(row=i, column=9).value = response_time(room_id)
        wb.save('%s.xlsx' % location)
print('Row %d has completed' % row2)
